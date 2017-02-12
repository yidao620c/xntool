#!/usr/bin/env python
# -*- encoding: utf-8 -*-
"""
openpyxl
The recommended package for reading and writing Excel 2010 files (ie: .xlsx)

xlsxwriter
An alternative package for writing data, formatting information and,
in particular, charts in the Excel 2010 format (ie: .xlsx)

通过一个schema.sql来生成excel表格的数据库设计文档

schema.sql示例：

-- -----------------------贷快发二期数据库-------------------------------
-- 公共用户表
DROP TABLE IF EXISTS t_public_user;
CREATE TABLE t_public_user (
  id                BIGINT PRIMARY KEY AUTO_INCREMENT COMMENT '主键ID',
  user_type         INTEGER COMMENT '用户类型 1:企业用户 2:个人用户',
  account           VARCHAR(20) NOT NULL COMMENT '账号',
  password          VARCHAR(32) NOT NULL COMMENT '密码',
  created_time      DATETIME DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
  updated_time      DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间'
) ENGINE=InnoDB AUTO_INCREMENT=1 DEFAULT CHARSET=utf8 COMMENT '公共用户表';

-- 贷款申请表
DROP TABLE IF EXISTS t_apply;
CREATE TABLE t_apply (
  id                BIGINT PRIMARY KEY AUTO_INCREMENT COMMENT '主键ID',
  created_time      DATETIME DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
  updated_time      DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间'
) ENGINE=InnoDB AUTO_INCREMENT=1 DEFAULT CHARSET=utf8 COMMENT '贷款申请表';

"""
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.styles import colors, borders, fills
import re


def load_schema(filename):
    """先加载schema.sql文件来获取所有建表语句"""
    result = []
    pat = re.compile(r'.* DEFAULT (\S+) .*')
    with open(filename, encoding='utf-8') as sqlfile:
        each_table = []  # 每张表定义
        temp_comment = ''
        for line in sqlfile:
            if line.startswith('--'):
                temp_comment = line.split('--')[1].strip()
            elif 'DROP TABLE' in line:
                each_table.insert(0, temp_comment)
                each_table.insert(1, line.strip().split()[-1][:-1])
            elif ' COMMENT ' in line and 'ENGINE=' not in line:
                col_arr = line.split()
                col_name = col_arr[0]
                col_type = col_arr[1]
                if 'PRIMARY KEY' in line or 'NOT NULL' in line:
                    col_null = 'NOT NULL'
                else:
                    col_null = ''
                col_remark = line.split(' COMMENT ')
                cr = col_remark[-1].strip().replace("'", "")
                defaultmatch = pat.match(line)
                default = defaultmatch.group(1) if defaultmatch else ''
                each_table.append((col_name, col_type, col_null,
                                   default, cr[:-1] if cr.endswith(',') else cr))
            elif 'ENGINE=' in line:
                # 单个表定义结束
                result.append(list(each_table))
                each_table.clear()
    return result


def write_dest(xlsx_name, schema_name):
    border = Border(
        left=Side(border_style=borders.BORDER_THIN, color='FF000000'),
        right=Side(border_style=borders.BORDER_THIN, color='FF000000'),
        top=Side(border_style=borders.BORDER_THIN, color='FF000000'),
        bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')
    )
    alignment = Alignment(horizontal='justify', vertical='bottom',
                          text_rotation=0, wrap_text=False,
                          shrink_to_fit=True, indent=0)
    fill = PatternFill(fill_type=None, start_color='FFFFFFFF')
    # 基本的样式
    basic_style = NamedStyle(font=Font(name='Microsoft YaHei')
                        , border=border, alignment=alignment
                        , fill=fill)
    title_style = basic_style.copy(
        font=Font(name='Microsoft YaHei', b=True, size=20, color='00215757'),
        alignment=Alignment(horizontal='center', vertical='bottom',
                            text_rotation=0, wrap_text=False,
                            shrink_to_fit=True, indent=0),
        fill=PatternFill(fill_type=fills.FILL_SOLID, start_color='00B2CBED'))
    header_style = basic_style.copy(
        font=Font(name='Microsoft YaHei', b=True, size=15, color='00215757'),
        fill=PatternFill(fill_type=fills.FILL_SOLID, start_color='00BAA87F'))
    common_style = basic_style.copy()
    link_style = basic_style.copy(font=Font(
        name='Microsoft YaHei', color=colors.BLUE, underline='single'))
    table_data = load_schema(schema_name)
    wb = Workbook()
    wb.active.title = "首页列表"

    for table in table_data:
        ws = wb.create_sheet(title=table[0])
        ws.merge_cells('E3:I3')  # 合并单元格
        ws['E3'].style = title_style
        ws['F2'].style = NamedStyle(border=Border(
            bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')))
        ws['G2'].style = NamedStyle(border=Border(
            bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')))
        ws['H2'].style = NamedStyle(border=Border(
            bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')))
        ws['I2'].style = NamedStyle(border=Border(
            bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')))
        ws['J3'].style = NamedStyle(border=Border(
            left=Side(border_style=borders.BORDER_THIN, color='FF000000')))
        ws['E3'] = table[0]
        ws['E4'].style = header_style
        ws['E4'] = '列名'
        ws['F4'].style = header_style
        ws['F4'] = '类型'
        ws['G4'].style = header_style
        ws['G4'] = '空值约束'
        ws['H4'].style = header_style
        ws['H4'] = '默认值'
        ws['I4'].style = header_style
        ws['I4'] = '备注'
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 40
        for idx, each_column in enumerate(table[2:]):
            ws['E{}'.format(idx + 5)].style = common_style
            ws['E{}'.format(idx + 5)] = each_column[0]
            ws['F{}'.format(idx + 5)].style = common_style
            ws['F{}'.format(idx + 5)] = each_column[1]
            ws['G{}'.format(idx + 5)].style = common_style
            ws['G{}'.format(idx + 5)] = each_column[2]
            ws['H{}'.format(idx + 5)].style = common_style
            ws['H{}'.format(idx + 5)] = each_column[3]
            ws['I{}'.format(idx + 5)].style = common_style
            ws['I{}'.format(idx + 5)] = each_column[4]
    ws = wb['首页列表']
    ws.merge_cells('D3:F3')
    ws['D3'].style = title_style
    ws['E2'].style = NamedStyle(border=Border(
        bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')))
    ws['F2'].style = NamedStyle(border=Border(
        bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')))
    ws['G3'].style = NamedStyle(border=Border(
        left=Side(border_style=borders.BORDER_THIN, color='FF000000')))
    ws['D3'] = 'MySQL数据库系统表'
    ws['D4'].style = header_style
    ws['D4'] = '编号'
    ws['E4'].style = header_style
    ws['E4'] = '表名'
    ws['F4'].style = header_style
    ws['F4'] = '详情链接'
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 35
    for inx, val in enumerate(table_data):
        ws['D{}'.format(inx + 5)].style = common_style
        ws['D{}'.format(inx + 5)] = inx + 1
        ws['E{}'.format(inx + 5)].style = common_style
        ws['E{}'.format(inx + 5)] = val[1]
        linkcell = ws['F{}'.format(inx + 5)]
        linkcell.style = link_style
        linkcell.value = val[0]
        linkcell.hyperlink = '#{0}!{1}'.format(val[0], 'E3')
    wb.save(filename=xlsx_name)


def main(dest_file, schema_file):
    write_dest(dest_file, schema_file)
